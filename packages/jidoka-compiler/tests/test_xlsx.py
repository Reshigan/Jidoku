import io, unittest
from openpyxl import Workbook
from jidoka_compiler.xlsx import compile_xlsx
from jidoka_core.ir import load_ir, IRValidationError

HEADER = ["object", "external_code", "tier", "depends_on", "field:name", "field:unit", "dp:floor"]


def _book() -> io.BytesIO:
    """Sheet 'Config' follows the convention; 'Notes' does not and must be skipped."""
    wb = Workbook()
    ws = wb.active; ws.title = "Config"
    ws.append(HEADER)
    ws.append(["TimeType", "ANN_LEAVE_ZAF", "A", "TimeAccountType:ANN_ACC_ZAF", "Annual Leave", "DAYS", ""])
    ws.append(["TimeAccountType", "ANN_ACC_ZAF", "B", "", "Annual Accrual", "", "DP-B11"])
    notes = wb.create_sheet("Notes"); notes.append(["free text", "no header convention"])
    second = wb.create_sheet("Payroll"); second.append(HEADER)
    second.append(["PayComponent", "1000", "C", "", "Basic", "ZAR", ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _signed():
    return compile_xlsx(_book(), "SuccessFactors", "SF-DEV", "komatsu.xlsx",
                        signed_by="A. Client", date="2026-01-05")


class TestXlsx(unittest.TestCase):
    def test_cell_range_provenance_is_exact(self):
        recs = _signed()
        self.assertEqual(recs[0]["source"]["cell_range"], "Config!A2:G2")
        self.assertEqual(recs[0]["source"]["sheet"], "Config")
        self.assertEqual(recs[0]["source"]["workbook"], "komatsu.xlsx")
        self.assertEqual(recs[1]["source"]["cell_range"], "Config!A3:G3")

    def test_multi_sheet_compiled_and_non_conforming_skipped(self):
        recs = _signed()
        self.assertEqual([r["source"]["sheet"] for r in recs], ["Config", "Config", "Payroll"])
        self.assertEqual(recs[2]["source"]["cell_range"], "Payroll!A2:G2")

    def test_blank_cell_becomes_decision_point_not_a_value(self):
        recs = _signed()
        unit = recs[1]["intent"]["unit"]
        self.assertIsNone(unit["value"])
        self.assertEqual(unit["decision_point"], "DP-GAP-CONFIG-F3")
        self.assertEqual(recs[1]["intent"]["floor"], {"value": None, "decision_point": "DP-B11"})
        _, dps = load_ir(recs)
        self.assertIn("SuccessFactors:TimeAccountType:ANN_ACC_ZAF", dps)

    def test_signed_workbook_loads(self):
        records, _ = load_ir(_signed())
        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["depends_on"] if isinstance(records[0], dict) else records[0].depends_on,
                         ["TimeAccountType:ANN_ACC_ZAF"])

    def test_unsigned_workbook_is_unloadable(self):
        recs = compile_xlsx(_book(), "SuccessFactors", "SF-DEV", "unsigned.xlsx")
        self.assertNotIn("signed_by", recs[0]["source"])
        with self.assertRaises(IRValidationError):
            load_ir(recs)
