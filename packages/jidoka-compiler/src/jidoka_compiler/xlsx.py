"""XLSX workbook -> IR with cell-level provenance. Same column convention as compile_csv.
Rule: the compiler NEVER guesses — a blank required cell becomes a decision_point, not a value.
Sign-off is the caller's: no signed_by => source lacks it => jidoka_core.ir refuses to load it."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REQUIRED_COLS = ("object", "external_code", "tier")


def _header(ws) -> dict[str, int] | None:
    """Sheets without the convention are skipped, not guessed at."""
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    cols = {str(v).strip(): i for i, v in enumerate(row, start=1) if v is not None}
    return cols if all(c in cols for c in REQUIRED_COLS) else None


def _blank_dp(sheet: str, col: str, row: int) -> str:
    """Deterministic dp_id for a cell the workbook left empty."""
    return f"DP-GAP-{sheet}-{col}{row}".upper().replace(" ", "_")


def compile_xlsx(path_or_stream, product: str, system_binding: str, workbook: str,
                 signed_by: str | None = None, date: str | None = None) -> list[dict]:
    """Compile every sheet carrying the header convention. Provenance is 'Sheet!A2:F2' per record."""
    wb = load_workbook(path_or_stream, data_only=True, read_only=True)
    out: list[dict] = []
    for ws in wb.worksheets:
        cols = _header(ws)
        if not cols:
            continue
        first, last = min(cols.values()), max(cols.values())
        span = f"{get_column_letter(first)}{{r}}:{get_column_letter(last)}{{r}}"
        for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cell = {name: ("" if i > len(row) or row[i - 1] is None else str(row[i - 1]).strip())
                    for name, i in cols.items()}
            if not any(cell.values()):
                continue  # blank spacer row, not a record
            intent: dict = {"externalCode": cell["external_code"]}
            for name, i in cols.items():
                val = cell[name]
                if name.startswith("field:"):
                    key = name[6:]
                    intent[key] = val if val else {
                        "value": None, "decision_point": _blank_dp(ws.title, get_column_letter(i), r)}
                elif name.startswith("dp:") and val:
                    intent[name[3:]] = {"value": None, "decision_point": val}
            source = {"workbook": workbook, "sheet": ws.title,
                      "cell_range": f"{ws.title}!{span.format(r=r)}"}
            if signed_by:
                source["signed_by"] = signed_by
            if date:
                source["date"] = date
            rec = {"object": cell["object"], "product": product, "system_binding": system_binding,
                   "tier": cell["tier"], "external_code": cell["external_code"],
                   "depends_on": [d for d in cell.get("depends_on", "").split("|") if d],
                   "intent": intent, "source": source}
            for req in REQUIRED_COLS:
                if not cell[req]:  # can't guess an object/tier/code — raise it to a human
                    intent.setdefault("_gaps", []).append(
                        {"value": None, "decision_point": _blank_dp(ws.title, get_column_letter(cols[req]), r),
                         "field": req})
            out.append(rec)
    wb.close()
    return out
